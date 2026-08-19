# -*- coding: utf-8 -*-
import os

import openpyxl
from openpyxl.styles import Font, PatternFill

import config
import db
import utils
from utils import log


def marcar_novos(itens: list[dict], fontes_confiaveis: set[str] | None = None,
                 execucao_id: int | None = None) -> list[dict]:
    """Grava a execução no SQLite (db.py) e devolve os itens com o
    campo 'novo': True/False já preenchido.

    fontes_confiaveis limita quais imóveis podem ser marcados como ausentes
    nesta rodada -- ver P-04 em db.salvar_execucao."""
    return db.salvar_execucao(itens, fontes_confiaveis, execucao_id)


def gerar_excel(itens: list[dict]) -> str:
    os.makedirs(config.PASTA_SAIDA, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Apartamentos"

    colunas = ["Novo?", "Fontes", "Nº fontes", "Título", "Endereço", "Bairro",
               "Aluguel", "Condomínio", "IPTU", "Custo total", "Custo completo?",
               "R$/m²", "Quartos", "Banheiros", "Andar", "Área (m²)",
               "Dias no mercado", "Visto 1ª vez em", "Link"]
    ws.append(colunas)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    fill_novo = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    # Novos primeiro, depois por preço crescente
    itens_ordenados = sorted(itens, key=lambda x: (not x.get("novo", False), x.get("preco") or 0))

    # Campo em branco é ambíguo na planilha: não dá para saber se o dado
    # não existe ou se o scraper não achou. "Não Localizado" torna a lacuna
    # explícita -- e é o que permite priorizar o levantamento manual.
    nl = utils.ou_nao_localizado

    for item in itens_ordenados:
        sites = item.get("sites") or ([item["site"]] if item.get("site") else [])
        urls = item.get("urls") or ([item["url"]] if item.get("url") else [])
        row = [
            "SIM" if item.get("novo") else "",
            ", ".join(sites),
            item.get("qtd_fontes", len(sites) or 1),
            item.get("titulo", ""),
            nl(item.get("logradouro")),
            nl(item.get("bairro")),
            nl(item.get("aluguel")),
            nl(item.get("condominio")),
            nl(item.get("iptu")),
            nl(item.get("custo_mensal_total") or item.get("preco")),
            # sem o condomínio o total é um piso, não o custo real
            "SIM" if item.get("custo_completo") else "NÃO",
            nl(item.get("preco_m2")),
            nl(item.get("quartos")),
            nl(item.get("banheiros")),
            nl(item.get("andar")),
            nl(item.get("area_m2")),
            nl(item.get("dias_anunciado")),
            item.get("primeiro_visto", ""),
            urls[0] if urls else "",
        ]
        ws.append(row)
        if item.get("novo"):
            for cell in ws[ws.max_row]:
                cell.fill = fill_novo

    for col, largura in zip("ABCDEFGHIJKLMNOPQRS",
                            [8, 30, 10, 42, 28, 18, 12, 12, 10, 13, 15, 10,
                             10, 11, 8, 12, 16, 16, 60]):
        ws.column_dimensions[col].width = largura

    wb.save(config.ARQUIVO_EXCEL)
    log.info(f"Excel salvo em {config.ARQUIVO_EXCEL}")
    return config.ARQUIVO_EXCEL
